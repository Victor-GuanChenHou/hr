import pyodbc
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime
import os
ENV = './.env' 
LOG_DIR='logs'
load_dotenv(os.path.join(os.getcwd(), ".env"))
def get_user_info(username):
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid=os.environ.get('HRDB_uid')
    HRDB_name=os.environ.get('HRDB_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
        "Trusted_Connection=no;"
            
    )
    cursor = conn.cursor()
    # SUBSTRING(UIDENTID, 2, LEN(UIDENTID) - 1) AS UIDENTID 身分證後九碼
    cursor.execute("SELECT EMPID, SUBSTRING(UIDENTID, 2, LEN(UIDENTID) - 1) AS UIDENTID, HECNAME ,DEPT_NO ,INADATE,CLASS FROM HRM.dbo.HRUSER WHERE STATE='A' AND EMPID = ?", (username,))
    row = cursor.fetchone()
    

    if row:
        cursor.execute("SELECT DEP_NAME,DEP_KIND FROM HRM.dbo.HRUSER_DEPT_BAS WHERE DEP_NO = ?", (row[3],))
        dep_row = cursor.fetchone()
        conn.close()

        if dep_row:
            return {'username': row[0], 'password': row[1], 'name': row[2], 'DEPT_NO':row[3],'INADATE':row[4],'CLASS':row[5],'DEPT_NAME':dep_row[0],'DEPT_KIND':dep_row[1]}
        else:
            return {'username': row[0], 'password': row[1], 'name': row[2], 'DEPT_NO':row[3],'INADATE':row[4],'CLASS':row[5],'DEPT_NAME':'NOT FOUND','DEPT_KIND':'NOT FOUND'}
    else:
        return None
def read_excel_compatible(filepath):
    try:
        df = pd.read_excel(filepath, engine='openpyxl')
        return df
    except Exception as e:
        print(f"pandas 讀取失敗，改用 openpyxl 處理: {e}")
        try:
            wb = load_workbook(filepath, data_only=True)

            # ✅ 確認是否有工作表
            if not wb.sheetnames:
                raise ValueError("Excel 檔案中沒有任何工作表")

            sheet = wb.active
            if sheet is None:
                raise ValueError("無法取得有效的工作表")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                raise ValueError("Excel 資料為空，或無有效資料")

            headers = list(rows[0])
            data = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
            return pd.DataFrame(data)

        except Exception as e2:
            raise RuntimeError(f"openpyxl 解析也失敗: {e2}")
def loglogin(username,ip):
    if not os.path.exists(LOG_DIR):
        os.makedirs(LOG_DIR)
    today = datetime.today().strftime('%Y-%m-%d')
    log_file = os.path.join(LOG_DIR, f'{today}.log')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_line = f'{now} | 使用者 {username} 從 {ip} 登入\n'
    
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(log_line)
def find_deptchie(username):
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid=os.environ.get('HRDB_uid')
    HRDB_name=os.environ.get('HRDB_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
        "Trusted_Connection=no;"
            
        # TrustServerCertificate=yes
    )
    cursor = conn.cursor()
    cursor.execute("SELECT DEPT_NO FROM HRM.dbo.HRUSER WHERE EMPID = ?", (username,))
    result = cursor.fetchone()
    if result[0]=='193':#判斷是否是央廚(央廚特別處理不抓上司MAIL)
        result='dcz01@kingza.com.tw'
        return result
    else:
        cursor.execute("""
            SELECT CHIEF.EMAIL
            FROM HRM.dbo.HRUSER EMP
            JOIN HRM.dbo.HRUSER_DEPT_BAS DEPT
                ON EMP.DEPT_NO = DEPT.DEP_NO
            JOIN HRM.dbo.HRUSER CHIEF
                ON DEPT.DEP_CHIEF = CHIEF.EMPID
            WHERE EMP.EMPID = ?
        """, (username,))
        result = cursor.fetchone()
    cursor.close()
    conn.close()

    if result:
        return result[0]
    else:
        return None
def get_dep_order(dep_name):
    if dep_name == '杏子豬排營運部':
        return 1
    elif dep_name.startswith('杏子'):
        return 2
    elif dep_name == '段純貞營運部':
        return 3
    elif dep_name.startswith('段純貞'):
        return 4
    elif dep_name.startswith('王將營運'):
        return 5
    elif dep_name.startswith('王將'):
        return 6
    elif dep_name.startswith('京都勝牛營運部'):
        return 7
    elif dep_name.startswith('勝牛'):
        return 8
    elif dep_name.startswith('橋村營運'):
        return 9
    elif dep_name.startswith('橋村'):
        return 10
    else:
        return 99
def docxuser():
    # 取得連線參數
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid = os.environ.get('HRDB_uid')
    HRDB_name = os.environ.get('HRDB_name')

    # 建立資料庫連線
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        f'SERVER={HRDB_host};'
        f'DATABASE={HRDB_name};'
        f'UID={HRDB_uid};'
        f'PWD={HRDB_password};'
    )
    cursor = conn.cursor()
    # SQL 查詢：取得在職 Class D 員工對應單位與身份別
    query_classd = """
    SELECT 
        D.DEP_NAME AS 單位名稱,
        U.EMPID AS 員工編號,
        U.HECNAME AS 員工姓名,
        T.UTNAME AS 身份別
    FROM HRM.dbo.HRUSER U
    LEFT JOIN HRM.dbo.HRUSER_DEPT_BAS D ON U.DEPT_NO = D.DEP_NO
    LEFT JOIN HRM.dbo.USERTYPE T ON U.UTYPE = T.UTYPE
    WHERE U.STATE = 'A' AND U.Class = 'D'
    """
    cursor.execute(query_classd)
    columns = [column[0] for column in cursor.description]

    # 取得所有資料
    rows = cursor.fetchall()

    # 轉成 DataFrame
    df = pd.DataFrame.from_records(rows, columns=columns)
    # 執行查詢
    #df = pd.read_sql(query_classd, conn)

    # 關閉連線
    conn.close()
     # ➕ 加入排序欄位
    df['dep_order'] = df['單位名稱'].apply(get_dep_order)

    # 🔽 排序
    df = df.sort_values(by=['dep_order', '單位名稱', '員工編號'])

    # 移除排序用欄位再轉成 list[dict]
    df = df.drop(columns=['dep_order'])

    return df.to_dict(orient='records')
def docxuser_END():
    year=datetime.today().year
    # 取得連線參數
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid = os.environ.get('HRDB_uid')
    HRDB_name = os.environ.get('HRDB_name')

    # 建立資料庫連線
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        f'SERVER={HRDB_host};'
        f'DATABASE={HRDB_name};'
        f'UID={HRDB_uid};'
        f'PWD={HRDB_password};'
    )
    cursor = conn.cursor()
    # SQL 查詢：取得在職 Class D 員工對應單位與身份別
    query_classd = """
    SELECT 
        D.DEP_NAME AS 單位名稱,
        U.EMPID AS 員工編號,
        U.HECNAME AS 員工姓名,
        T.UTNAME AS 身份別
    FROM HRM.dbo.HRUSER U
    LEFT JOIN HRM.dbo.HRUSER_DEPT_BAS D ON U.DEPT_NO = D.DEP_NO
    LEFT JOIN HRM.dbo.USERTYPE T ON U.UTYPE = T.UTYPE
    WHERE (U.STATE = 'A' OR U.QUITDATE LIKE '{year}%') AND U.Class = 'D'
    """
    cursor.execute(query_classd)
    columns = [column[0] for column in cursor.description]

    # 取得所有資料
    rows = cursor.fetchall()

    # 轉成 DataFrame
    df = pd.DataFrame.from_records(rows, columns=columns)
    # 執行查詢
    #df = pd.read_sql(query_classd, conn)

    # 關閉連線
    conn.close()
     # ➕ 加入排序欄位
    df['dep_order'] = df['單位名稱'].apply(get_dep_order)

    # 🔽 排序
    df = df.sort_values(by=['dep_order', '單位名稱', '員工編號'])

    # 移除排序用欄位再轉成 list[dict]
    df = df.drop(columns=['dep_order'])

    return df.to_dict(orient='records')

def exe_get_holidaydata():
    import pyodbc
    import pandas as pd
    import os
    from dotenv import load_dotenv
    from datetime import datetime, timedelta
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)
    ENV = './.env' 
    load_dotenv(os.path.join(os.getcwd(), ".env"))


    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid=os.environ.get('HRDB_uid')
    HRDB_name=os.environ.get('HRDB_name')
    # 資料庫連線設定
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
    )
    #排序
    def get_dep_order(dep_name):
        if not isinstance(dep_name, str):
            return 99
        if dep_name == '杏子豬排營運部':
            return 1
        elif dep_name.startswith('杏子'):
            return 2
        elif dep_name == '段純貞營運部':
            return 3
        elif dep_name.startswith('段純貞'):
            return 4
        elif dep_name.startswith('王將營運'):
            return 5
        elif dep_name.startswith('王將'):
            return 6
        elif dep_name.startswith('京都勝牛營運部'):
            return 7
        elif dep_name.startswith('勝牛'):
            return 8
        elif dep_name.startswith('橋村營運'):
            return 9
        elif dep_name.startswith('橋村'):
            return 10
        elif dep_name.startswith('雞三和營運部'):
            return 11
        elif dep_name.startswith('雞三和'):
            return 12
        else:
            return 99
    # 取得今天的日期
    today = datetime.today()

    # 找出「本月的第一天」再減一天，就會是「上個月的最後一天」
    first_day_this_month = today.replace(day=1)
    last_day_last_month = first_day_this_month - timedelta(days=1)

    # 轉換成 YYMM 格式
    yymm_last_month = last_day_last_month.strftime("%Y%m")
    # 1. 查詢 CLASSDA
    query_classda = f"""
    SELECT CPNYID, CLASSDA, EMPID, CLASS
    FROM HRM.dbo.CLASSDA
    WHERE CLASSDA LIKE'{yymm_last_month}%' AND CLASS='H'
    """
    df_classda = pd.read_sql(query_classda, conn)

    if df_classda.empty:
        print("查無 CLASSDA 資料")
    else:
        # 2. 查詢 HRUSER 中的 EMPID 對應 DEPT_NO, HECNAME, UTYPE
        emp_ids = tuple(df_classda['EMPID'].unique())
        query_hruser = f"""
        SELECT EMPID, DEPT_NO, HECNAME, UTYPE ,STATE ,UIDENTID
        FROM HRM.dbo.HRUSER_{yymm_last_month}
        WHERE EMPID IN {emp_ids}   AND (UTYPE='F' OR UTYPE='H')
        """
        df_hruser = pd.read_sql(query_hruser, conn)
        emp_ids_c = tuple(df_hruser[df_hruser['STATE'] == 'C']['EMPID'].unique())
        query_hruser = f"""
        SELECT EMPID, DEPT_NO, HECNAME, UTYPE ,STATE ,UIDENTID
        FROM HRM.dbo.HRUSER_{yymm_last_month}
        WHERE EMPID IN {emp_ids}   AND (UTYPE='F' OR UTYPE='H')
        """
        df_hruser = pd.read_sql(query_hruser, conn)
        UIDENTID_c = tuple(df_hruser[df_hruser['STATE'] == 'C']['UIDENTID'].unique())

        
        if UIDENTID_c:
            #查 HRUSER（不加月份）找 UIDENTID 有 STATE='A' 的
            query_current = f"""
            SELECT EMPID, DEPT_NO, HECNAME, UTYPE ,STATE ,UIDENTID
            FROM HRM.dbo.HRUSER
            WHERE UIDENTID IN {UIDENTID_c} AND STATE='A' 
            """
            df_current = pd.read_sql(query_current, conn)
            
            # 過濾 C，只保留有對應 A 的
            valid_uid = set(df_current['UIDENTID'])
            df_hruser = df_hruser[~((df_hruser['STATE'] == 'C') & (~df_hruser['UIDENTID'].isin(valid_uid)))]


    
        
        active_emp_ids = df_hruser['EMPID'].unique()
        df_classda = df_classda[df_classda['EMPID'].isin(active_emp_ids)]
        # 3. 合併 CLASSDA + HRUSER
        df_merged = pd.merge(df_classda, df_hruser, on='EMPID', how='left')

        # 4. 查詢 HRUSER_DEPT_BAS 取得 DEP_NAME 和 DEP_CHIEF
        dept_nos = tuple(df_merged['DEPT_NO'].dropna().unique())
        query_dept = f"""
        SELECT DEP_NO, DEP_NAME, DEP_CHIEF
        FROM HRM.dbo.HRUSER_DEPT_BAS
        WHERE DEP_NO IN {dept_nos}
        """
        df_dept = pd.read_sql(query_dept, conn)
        df_merged['DEPT_NO'] = df_merged['DEPT_NO'].astype(str)
        df_dept['DEP_NO'] = df_dept['DEP_NO'].astype(str)

        # 合併取得 DEP_NAME 和 DEP_CHIEF
        df_merged = pd.merge(df_merged, df_dept, left_on='DEPT_NO', right_on='DEP_NO', how='left')

        # 查詢主管 EMAIL
        chief_ids = tuple(df_merged['DEP_CHIEF'].dropna().unique())
        query_chief = f"""
        SELECT EMPID, EMAIL
        FROM HRM.dbo.HRUSER
        WHERE EMPID IN {chief_ids}
        """
        df_chief = pd.read_sql(query_chief, conn)
    # df_chief.rename(columns={'EMPID': 'DEP_CHIEF', 'EMAIL': '主管'}, inplace=True)

        # 合併主管 EMAIL
        df_dept = pd.merge(df_dept, df_chief, left_on='DEP_CHIEF', right_on='EMPID', how='left')
        df_dept = df_dept.rename(columns={'EMAIL': '主管'})

        # 5-3. 特殊處理「段純貞中央工廠」的主管 Email
        df_dept.loc[df_dept['DEP_NAME'] == '段純貞中央工廠', '主管'] = 'dcz01@kingza.com.tw'

        # 合併進 df_merged
        df_merged = pd.merge(df_merged, df_dept[['DEP_NO', '主管']], on='DEP_NO', how='left')
            
        
        # 5. 查詢 USERTYPE 取得身份別名稱
        utypes = tuple(df_merged['UTYPE'].dropna().unique())
        query_utype = f"""
        SELECT UTYPE, UTNAME
        FROM HRM.dbo.USERTYPE
        WHERE UTYPE IN {utypes}
        """
        df_utype = pd.read_sql(query_utype, conn)
        df_merged = pd.merge(df_merged, df_utype, on='UTYPE', how='left')
        # 6. 查詢 CLASSSET 取得班別名稱 CLNAME
        classes = tuple(df_merged['CLASS'].dropna().unique())
        query_classset = f"""
        SELECT CLASS, CLNAME
        FROM HRM.dbo.CLASSSET
        WHERE CLASS = 'H'
        """
        df_classset = pd.read_sql(query_classset, conn)

        df_merged = pd.merge(df_merged, df_classset, on='CLASS', how='left')
        # 7. 最終欄位輸出
        df_result = df_merged[['DEP_NAME', 'EMPID', 'HECNAME', 'CLASSDA','CLNAME', 'UTNAME', '主管']]

        # 改欄位名稱為中文
        df_result = df_result.rename(columns={
            'DEP_NAME': '單位名稱',
            'EMPID': '員工編號',
            'HECNAME': '員工姓名',
            'CLASSDA': '日期',
            'CLNAME' :'班別',
            'UTNAME': '身份別',
            '主管': '主管'
        })

        # 排序
        df_result['DEP_ORDER'] = df_result['單位名稱'].apply(get_dep_order)
        df_result = df_result.sort_values(by=['DEP_ORDER', '單位名稱', '員工編號'])
        df_result = df_result.drop(columns=['DEP_ORDER'])

        # 印出結果
        output_filename = './uploads/upload_month/'+yymm_last_month+'_國定假日.xlsx'  # 你可以動態產生名稱，或固定名稱

        df_result.to_excel(output_filename, index=False)

        print(f"已成功輸出 Excel 檔案：{output_filename}")

        logdata='success'
    return logdata    
def get_dept_people(depts):
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid=os.environ.get('HRDB_uid')
    HRDB_name=os.environ.get('HRDB_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
        "Trusted_Connection=no;"
            
    )
    placeholders = ",".join("?" for _ in depts)

    sql = f"""
    SELECT DEP_NAME, DEP_KIND ,DEP_NO
    FROM HRM.dbo.HRUSER_DEPT_BAS
    WHERE CPNYID='42756204' AND DEP_NAME IN ({placeholders})
    """
    cursor = conn.cursor()
    cursor.execute(sql, depts)
    rows = cursor.fetchall()
    data=[]
    for i in range(len(rows)):
        cursor.execute("SELECT EMPID, HECNAME ,DEPT_NO ,INADATE FROM HRM.dbo.HRUSER WHERE STATE='A' AND DEPT_NO = ?", (rows[i][2],))
        empid = cursor.fetchall()
        if empid:
            for z in range(len(empid)):
                add={'dept_no':rows[i][2],'dept_name':rows[i][0],'username':empid[z][0],'name':empid[z][1],'dates':empid[z][3]}
                data.append(add)
    return(data)
def getall_empid():
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid=os.environ.get('HRDB_uid')
    HRDB_name=os.environ.get('HRDB_name')
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={HRDB_host};"
        f"DATABASE={HRDB_name};"
        f"UID={HRDB_uid};"
        f"PWD={HRDB_password};"
        "Trusted_Connection=no;"
            
    )
    cursor = conn.cursor()
    # SUBSTRING(UIDENTID, 2, LEN(UIDENTID) - 1) AS UIDENTID 身分證後九碼
    cursor.execute("""
        SELECT 
            U.EMPID,
            U.HECNAME,
            U.DEPT_NO,
            U.INADATE,
            U.CLASS,
            D.DEP_NAME,
            D.DEP_KIND
        FROM HRM.dbo.HRUSER AS U
        LEFT JOIN HRM.dbo.HRUSER_DEPT_BAS AS D
            ON U.DEPT_NO = D.DEP_NO
        WHERE U.STATE = 'A'
        AND U.CPNYID = '42756204'
        AND U.DEPT_NO != ''
    """)

    rows = cursor.fetchall()

    alldata = []
    for row in rows:
        data = {
            'username': row[0],
            'name': row[1],
            'DEPT_NO': row[2],
            'INADATE': row[3],
            'CLASS': row[4],
            'dept_name': row[5] if row[5] else 'NOT FOUND',
            'dept_no': row[6] if row[6] else 'NOT FOUND'
        }
        alldata.append(data)

    conn.close()
    return alldata if alldata else None

def docxuser_manager_mail():
    # 取得連線參數
    HRDB_host = os.environ.get('HRDB_host')
    HRDB_password = os.environ.get('HRDB_password')
    HRDB_uid = os.environ.get('HRDB_uid')
    HRDB_name = os.environ.get('HRDB_name')

    # 建立資料庫連線
    conn = pyodbc.connect(
        'DRIVER={ODBC Driver 17 for SQL Server};'
        f'SERVER={HRDB_host};'
        f'DATABASE={HRDB_name};'
        f'UID={HRDB_uid};'
        f'PWD={HRDB_password};'
    )
    cursor = conn.cursor()
    # SQL 查詢：取得在職 Class D 員工對應單位與身份別
    query_classd = """
    SELECT 
        D.DEP_NAME AS 單位名稱,
        
        U.EMPID AS 員工編號,
        U.HECNAME AS 員工姓名,
        T.UTNAME AS 身份別,
        D.DEP_CHIEF 主管,
        CHIEF.EMAIL 單位主管信箱
    FROM HRM.dbo.HRUSER U
    LEFT JOIN HRM.dbo.HRUSER_DEPT_BAS D ON U.DEPT_NO = D.DEP_NO
    LEFT JOIN HRM.dbo.USERTYPE T ON U.UTYPE = T.UTYPE
    LEFT JOIN HRM.dbo.HRUSER CHIEF ON D.DEP_CHIEF = CHIEF.EMPID
    WHERE U.STATE = 'A' AND U.Class = 'D'
    """
    cursor.execute(query_classd)
    columns = [column[0] for column in cursor.description]

    # 取得所有資料
    rows = cursor.fetchall()

    # 轉成 DataFrame
    df = pd.DataFrame.from_records(rows, columns=columns)
    # 執行查詢
    #df = pd.read_sql(query_classd, conn)

    # 關閉連線
    conn.close()
     # ➕ 加入排序欄位
    df['dep_order'] = df['單位名稱'].apply(get_dep_order)

    # 🔽 排序
    df = df.sort_values(by=['dep_order', '單位名稱', '員工編號'])

    # 移除排序用欄位再轉成 list[dict]
    df = df.drop(columns=['dep_order'])

    return df.to_dict(orient='records')
